# -*- coding: utf-8 -*-
"""Genera sept_data.json para la pestana 4 (FCST Septiembre + Riesgo).

Fuentes:
  - "FCST Septiembre 2026.xlsx" (OneDrive, hoja Abarrotes): target .com por
    categoria (fila 33) y su comparable LY / Sept-2025 (fila 79).
  - cat_agg.csv: crecimiento YoY MTD/L7D real que ya trae el tablero.
  - merged_full.csv: catalogo item-level para armar accionables de septiembre.
  - reporte_internal_search_boosteos.html: terminos de busqueda prioritarios
    ya identificados (varios marcados Fiestas Patrias).
"""
import json
import os
import re
import pandas as pd

# merged_full.csv se lee UNA sola vez aqui arriba (item-level, ya trae
# YTD/MTD/L7D/AMX) -- todas las secciones de abajo (categorias, accionables
# de FP, evento AMX) reusan este mismo dataframe, no se vuelve a leer el CSV.
full = pd.read_csv('merged_full.csv')

# ---------- 1. Leer FCST Septiembre (target .com por categoria + comparable LY) ----------
# OJO: NO hardcodear la ruta completa -- la carpeta intermedia se ha
# movido antes sin avisar (de "OneDrive\W2\..." a "OneDrive\Seguros
# 2026\W2\..."). Se busca el archivo por NOMBRE con rglob desde la raiz
# del OneDrive para sobrevivir a que Alberto siga reorganizando carpetas
# -- esto corre TODOS LOS DIAS via Task Scheduler, no se puede dar el lujo
# de tronar por un folder renombrado.
import pathlib
onedrive_root = pathlib.Path(r"C:\Users\a0f07dn\OneDrive - Walmart Inc")
fcst_path = next(onedrive_root.rglob("FCST Septiembre 2026.xlsx"))

import openpyxl
wb = openpyxl.load_workbook(fcst_path, data_only=True)
ws = wb['Abarrotes']

# Columnas fila 33/79: 4=cat41, 6=cat46, 8=cat49, 10=cat53, 12=cat43, 14=cat68, 16=Total
COL_BY_CAT = {41: 4, 46: 6, 49: 8, 53: 10, 43: 12, 68: 14}
fcst_row = {c: ws.cell(row=33, column=col).value for c, col in COL_BY_CAT.items()}
fcst_total = ws.cell(row=33, column=16).value
ly_row = {c: ws.cell(row=79, column=col + 1).value for c, col in COL_BY_CAT.items()}
ly_total = sum(ly_row.values())

# ---------- 2. Crecimiento real YTD/MTD/L7D por categoria ----------
# El Gap de FCST usa YTD (no L7D) como base de tendencia -- decision de
# Alberto (01-sep-2026): L7D es muy ruidoso semana a semana para proyectar
# un mes completo, YTD da una base mas estable acumulada del anio.
cat_agg = pd.read_csv('cat_agg.csv').set_index('Cat_Nbr')
ytd_cat = full.groupby('Cat_Nbr')[['Com_Pesos_YTD', 'Com_Pesos_YTDLY']].sum()

categorias = []
for cat_nbr, fcst_val in fcst_row.items():
    cat_desc = cat_agg.loc[cat_nbr, 'Cat_Desc']
    ly_val = ly_row[cat_nbr]
    crec_mtd = cat_agg.loc[cat_nbr, 'Crec_Com_MTD']
    crec_l7d = cat_agg.loc[cat_nbr, 'Crec_Com_L7D']
    crec_ytd = (
        (ytd_cat.loc[cat_nbr, 'Com_Pesos_YTD'] - ytd_cat.loc[cat_nbr, 'Com_Pesos_YTDLY'])
        / ytd_cat.loc[cat_nbr, 'Com_Pesos_YTDLY']
    )
    com_mtd_actual = cat_agg.loc[cat_nbr, 'Com_Pesos_MTD']
    growth_needed = (fcst_val - ly_val) / ly_val
    trend_estimate = ly_val * (1 + crec_ytd)
    gap = trend_estimate - fcst_val
    gap_pct = gap / fcst_val
    if crec_ytd < growth_needed - 0.05:
        risk = 'Alto'
    elif crec_ytd < growth_needed:
        risk = 'Moderado'
    else:
        risk = 'Bajo'
    categorias.append({
        'cat_nbr': int(cat_nbr), 'cat_desc': cat_desc,
        'fcst_sept': round(fcst_val, 2), 'ly_sept': round(ly_val, 2),
        'growth_needed': round(growth_needed, 4),
        'crec_mtd_actual': round(float(crec_mtd), 4), 'crec_l7d_actual': round(float(crec_l7d), 4),
        'crec_ytd_actual': round(float(crec_ytd), 4),
        'com_mtd_actual': round(float(com_mtd_actual), 2),
        'trend_estimate': round(trend_estimate, 2), 'gap': round(gap, 2), 'gap_pct': round(gap_pct, 4),
        'risk': risk,
    })
categorias.sort(key=lambda c: c['fcst_sept'], reverse=True)

trend_total = sum(c['trend_estimate'] for c in categorias)

# Crecimiento MTD/L7D total real de .com Abarrotes -- se lee de dashboard_data.json
# (pestana 1), NUNCA se hardcodea aqui para que no se desactualice con el resto
# del tablero cuando se corre la rutina diaria. YTD si se calcula aqui mismo
# (pestana 1 no lo expone) sumando directo del item-level ya cargado en 'full'.
with open('dashboard_data.json', 'r', encoding='utf-8') as f:
    tab1_kpis = json.load(f)['kpis']

tot_com_ytd, tot_com_ytdly = float(full['Com_Pesos_YTD'].sum()), float(full['Com_Pesos_YTDLY'].sum())
crec_ytd_actual_total = (tot_com_ytd - tot_com_ytdly) / tot_com_ytdly

kpis = {
    'fcst_total': round(fcst_total, 2), 'ly_total': round(ly_total, 2),
    'growth_needed_total': round((fcst_total - ly_total) / ly_total, 4),
    'crec_mtd_actual_total': tab1_kpis['com_mtd_growth'],
    'crec_l7d_actual_total': tab1_kpis['com_l7d_growth'],
    'crec_ytd_actual_total': round(crec_ytd_actual_total, 4),
    'trend_total': round(trend_total, 2),
    'gap_total': round(trend_total - fcst_total, 2),
    'gap_pct_total': round((trend_total - fcst_total) / fcst_total, 4),
}

# ---------- 3. Item-level: subcategorias de Fiestas Patrias / boost de busqueda ----------

TARGET_SUBCATS = {
    ' MAYONESA (INDIVIDUAL)': ('Mayonesa (boost busqueda)', True),
    ' CAFE. SOLUBLE': ('Nescafe/Cafe (boost busqueda)', False),
    ' LACTEOS CULINARIOS': ('Media crema/Carnation/Clavel/Lechera (FP guisos)', True),
    ' MEZCLAS DE ESPECIAS Y SAZONADORES': ('Maggi/Knorr (FP guisos)', True),
    ' SALSAS PARA LA MESA': ('Salsa inglesa (FP)', True),
    ' LECHE. ENTERA': ('Leche entera (boost busqueda)', False),
    ' CEREAL': ('Cereal (FP)', True),
    ' UNTABLES': ('Nutella (boost busqueda)', False),
    ' CAFE. SUSTITUTO DE CREMA': ('Coffee-Mate (boost busqueda)', False),
    ' TOMATE': ('Pure de tomate (FP mole/guisos)', True),
    ' CALDOS. POLLO': ('Caldo de pollo (FP pozole)', True),
    ' SALSAS PARA COCINAR': ('Salsa (FP)', True),
    ' MOLE': ('Mole (FP alta prioridad)', True),
    ' ACEITE (INDIVIDUAL)': ('Aceite (FP)', True),
    ' GALLETAS': ('Galletas/Oreo/Saladitas (boost busqueda)', False),
    ' EDULCORANTES': ('Splenda (boost busqueda)', False),
    ' MODIFICADORES DE LECHE': ('Nesquik (boost busqueda)', False),
}
sub = full[full['Sub_Cat_Desc'].isin(TARGET_SUBCATS.keys())].copy()
sub['Boost_Motivo'] = sub['Sub_Cat_Desc'].map(lambda s: TARGET_SUBCATS[s][0])
sub['Fiestas_Patrias'] = sub['Sub_Cat_Desc'].map(lambda s: TARGET_SUBCATS[s][1])
sub['Share_Com_MTD_calc'] = sub['Com_Pesos_MTD'] / (sub['Com_Pesos_MTD'] + sub['Piso_Pesos_MTD'])


def item_row(r):
    return {
        'item_nbr': int(r['Item_Nbr']), 'item_desc': r['Item_Desc_1'], 'cat_desc': r['Cat_Desc'],
        'sub_cat_desc': r['Sub_Cat_Desc'].strip(), 'boost_motivo': r['Boost_Motivo'],
        'fiestas_patrias': bool(r['Fiestas_Patrias']),
        'piso_mtd': round(float(r['Piso_Pesos_MTD']), 2), 'com_mtd': round(float(r['Com_Pesos_MTD']), 2),
        'share_com_mtd': round(float(r['Share_Com_MTD_calc']), 4),
        'crec_com_mtd': (round(float(r['Crecimiento_Com_Pesos_MTD']), 4) if pd.notna(r['Crecimiento_Com_Pesos_MTD']) else None),
        'crec_com_l7d': (round(float(r['Crecimiento_Com_Pesos_L7D']), 4) if pd.notna(r['Crecimiento_Com_Pesos_L7D']) else None),
        'en_parrilla': bool(r['En_Parrilla']), 'promo_vigente': bool(r['Promo_Vigente']),
        'semaforo': r['Semaforo_OH'],
    }


# Apagar incendios: gran volumen fisico, .com cayendo fuerte en L7D.
apagar = sub[(sub['Piso_Pesos_MTD'] > 5_000_000) & (sub['Crecimiento_Com_Pesos_L7D'] < -0.20)]
apagar = apagar.sort_values('Piso_Pesos_MTD', ascending=False).head(10)

# Doblar apuesta: ya en parrilla+promo, momentum sostenido (MTD y L7D positivos).
doblar = sub[
    sub['En_Parrilla'] & sub['Promo_Vigente']
    & (sub['Crecimiento_Com_Pesos_MTD'] > 0.20) & (sub['Crecimiento_Com_Pesos_L7D'] > 0)
]
doblar = doblar.sort_values('Com_Pesos_MTD', ascending=False).head(12)

# Blanco total: demanda fisica real, sin promo, mix .com por debajo del promedio Abarrotes (~9.8%).
blanco = sub[(sub['Piso_Pesos_MTD'] > 2_000_000) & (~sub['Promo_Vigente']) & (sub['Share_Com_MTD_calc'] < 0.09)]
blanco = blanco.sort_values('Piso_Pesos_MTD', ascending=False).head(12)

items = {
    'apagar_incendios': [item_row(r) for _, r in apagar.iterrows()],
    'doblar_apuesta': [item_row(r) for _, r in doblar.iterrows()],
    'blanco_total': [item_row(r) for _, r in blanco.iterrows()],
}

# ---------- 4. Evento "A la Mexicana" (9-16 sep 2026) ----------
# Reusa 'full' (merged_full.csv) que ya trae las columnas *_AMX/*_AMXLY
# porque build_merge.py pasa TODAS las columnas de raw_bq_item_total.csv
# sin filtrar -- no hizo falta tocar build_merge.py/cat_agg.csv para esto.
from datetime import date

AMX_INI = date(2026, 9, 9)
AMX_FIN = date(2026, 9, 16)
hoy = pd.Timestamp.now().date()
amx_iniciado = hoy >= AMX_INI


def safe_growth(cur, prev):
    # Antes de que arranque el evento, Com_Pesos_AMX/Piso_Pesos_AMX son 0 de
    # verdad (todavia no hay ventas), pero *_AMXLY ya trae datos reales de
    # 2025 -- calcular el % ahi daria un falso "-100%" que asusta sin
    # decir nada util. Se regresa None ("n/d" en el front) hasta que el
    # evento arranque de verdad.
    if not amx_iniciado:
        return None
    if prev in (0, None) or pd.isna(prev):
        return None
    return (cur - prev) / prev


amx_cat = full.groupby(['Cat_Nbr', 'Cat_Desc']).agg(
    Com_Pesos_AMX=('Com_Pesos_AMX', 'sum'), Com_Pesos_AMXLY=('Com_Pesos_AMXLY', 'sum'),
    Piso_Pesos_AMX=('Piso_Pesos_AMX', 'sum'), Piso_Pesos_AMXLY=('Piso_Pesos_AMXLY', 'sum'),
).reset_index()

amx_categorias = []
for _, r in amx_cat.iterrows():
    com_amx, com_amxly = float(r['Com_Pesos_AMX']), float(r['Com_Pesos_AMXLY'])
    piso_amx, piso_amxly = float(r['Piso_Pesos_AMX']), float(r['Piso_Pesos_AMXLY'])
    total_amx = com_amx + piso_amx
    crec_com = safe_growth(com_amx, com_amxly)
    crec_piso = safe_growth(piso_amx, piso_amxly)
    amx_categorias.append({
        'cat_nbr': int(r['Cat_Nbr']), 'cat_desc': r['Cat_Desc'],
        'com_amx': round(com_amx, 2), 'com_amxly': round(com_amxly, 2),
        'crec_com_amx': (round(crec_com, 4) if crec_com is not None else None),
        'piso_amx': round(piso_amx, 2), 'piso_amxly': round(piso_amxly, 2),
        'crec_piso_amx': (round(crec_piso, 4) if crec_piso is not None else None),
        'share_com_amx': (round(com_amx / total_amx, 4) if total_amx > 0 else None),
    })
amx_categorias.sort(key=lambda c: c['com_amx'], reverse=True)

tot_com_amx, tot_com_amxly = float(full['Com_Pesos_AMX'].sum()), float(full['Com_Pesos_AMXLY'].sum())
tot_piso_amx, tot_piso_amxly = float(full['Piso_Pesos_AMX'].sum()), float(full['Piso_Pesos_AMXLY'].sum())
crec_com_amx_total = safe_growth(tot_com_amx, tot_com_amxly)
crec_piso_amx_total = safe_growth(tot_piso_amx, tot_piso_amxly)

dias_totales = (AMX_FIN - AMX_INI).days + 1
iniciado = amx_iniciado
terminado = hoy > AMX_FIN
if not iniciado:
    dias_transcurridos = 0
    amx_status_msg = f"El evento arranca el {AMX_INI.strftime('%d-%b-%Y')} -- estos valores se activan solos ese dia con la corrida diaria de siempre, no hace falta tocar nada."
elif not terminado:
    dias_transcurridos = (hoy - AMX_INI).days + 1
    amx_status_msg = f"Evento en curso: dia {dias_transcurridos} de {dias_totales} ({AMX_INI.strftime('%d-%b')} -> {AMX_FIN.strftime('%d-%b')})."
else:
    dias_transcurridos = dias_totales
    amx_status_msg = f"Evento cerrado ({AMX_INI.strftime('%d-%b')} -> {AMX_FIN.strftime('%d-%b')}) -- resultado final vs LY."

evento_amx = {
    'fecha_ini': AMX_INI.isoformat(), 'fecha_fin': AMX_FIN.isoformat(),
    'iniciado': iniciado, 'terminado': terminado,
    'dias_transcurridos': dias_transcurridos, 'dias_totales': dias_totales,
    'status_msg': amx_status_msg,
    'kpis': {
        'com_amx': round(tot_com_amx, 2), 'com_amxly': round(tot_com_amxly, 2),
        'crec_com_amx': (round(crec_com_amx_total, 4) if crec_com_amx_total is not None else None),
        'piso_amx': round(tot_piso_amx, 2), 'piso_amxly': round(tot_piso_amxly, 2),
        'crec_piso_amx': (round(crec_piso_amx_total, 4) if crec_piso_amx_total is not None else None),
    },
    'categorias': amx_categorias,
}

# ---------- 5. Contexto del reporte de boosteos de busqueda (terminos Fiestas Patrias) ----------
boost_html = open('../reporte_internal_search_boosteos.html', encoding='utf-8').read()
m = re.search(r'const DATA = (\{.*?\});', boost_html, re.DOTALL)
boost_data = json.loads(m.group(1))
sept_kw_ctx = {
    'fiestas_26_top': boost_data['sept_kw'].get('fiestas_26_top', []),
    'fiestas_26_count': boost_data['sept_kw'].get('fiestas_26_count'),
    'climbers': boost_data['sept_kw'].get('climbers', []),
}

# ---------- 5. Insights ejecutivos ----------
def fmt_pct(x):
    sign = '+' if x >= 0 else ''
    return f"{sign}{x*100:.1f}%"


riesgo_alto = [c['cat_desc'] for c in categorias if c['risk'] == 'Alto']
concentracion_txt = (
    f"<strong>[Concentracion] El riesgo esta concentrado en {', '.join(riesgo_alto)}</strong> "
    "&mdash; vienen desacelerando en su tendencia YTD justo cuando el FCST les pide mas crecimiento, no menos."
) if riesgo_alto else (
    "<strong>[Concentracion] Ninguna categoria esta en riesgo Alto</strong> con la tendencia YTD actual "
    "&mdash; el riesgo, si lo hay, es moderado y disperso entre categorias."
)

insights_top = [
    f"<strong>[FCST] El objetivo de septiembre es {fmt_pct(kpis['growth_needed_total'])} YoY</strong> vs Sept 2025 "
    f"(${ly_total/1e6:.1f}M &rarr; ${fcst_total/1e6:.1f}M) &mdash; el YTD real de .com Abarrotes va en {fmt_pct(kpis['crec_ytd_actual_total'])} "
    f"(la base que usa el estimado de abajo), con un MTD de agosto de {fmt_pct(kpis['crec_mtd_actual_total'])} y una ultima semana de {fmt_pct(kpis['crec_l7d_actual_total'])}.",
    (
        f"<strong>[Riesgo] Si el ritmo YTD se mantiene</strong> (base mas estable que una sola semana suelta), el estimado de septiembre sale en "
        f"${kpis['trend_total']/1e6:.1f}M &mdash; un faltante de ${abs(kpis['gap_total'])/1e6:.1f}M ({fmt_pct(kpis['gap_pct_total'])}) vs el target."
        if kpis['gap_total'] < 0 else
        f"<strong>[Positivo] Con el ritmo YTD acumulado</strong>, el estimado de tendencia (${kpis['trend_total']/1e6:.1f}M) ya supera el target."
    ),
    concentracion_txt,
    f"<strong>[Accionables] {len(items['apagar_incendios'])} items</strong> con alto volumen en Piso estan cayendo fuerte en .com "
    f"(candidatos a 'apagar incendio' antes de septiembre), <strong>{len(items['doblar_apuesta'])} items</strong> ya tienen momentum "
    f"sostenido para escalar, y <strong>{len(items['blanco_total'])} items</strong> tienen demanda fisica real pero cero promo vigente.",
]

data = {
    'generated_at': pd.Timestamp.now().strftime('%Y-%m-%d %H:%M'),
    'kpis': kpis,
    'categorias': categorias,
    'items': items,
    'evento_amx': evento_amx,
    'sept_kw_context': sept_kw_ctx,
    'insights_top': insights_top,
}

with open('sept_data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print('KPIs:', json.dumps(kpis, indent=2, ensure_ascii=False))
print('Categorias:', len(categorias))
print('Items:', {k: len(v) for k, v in items.items()})
print('Evento AMX:', json.dumps(evento_amx['kpis'], indent=2, ensure_ascii=False), '|', evento_amx['status_msg'])
print('Guardado sept_data.json')
